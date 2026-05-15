"""
Standalone Maintenance Dashboard — Flask backend.
Serves only the Maintenance page and its required API endpoints.
"""

from flask import Flask, jsonify, send_from_directory, request
import os
import re

# ── Service imports (all kept — Maintenance page depends on all of them) ────────
from maintenance_service import (
    build_maintenance_overview_payload,
    build_filter_payload,
    build_list_payload,
    build_monthly_payload,
    build_summary_payload,
    build_timeline_payload,
    build_equipment_filter_payload,
    build_equipment_list_payload,
    build_equipment_monthly_payload,
    build_equipment_summary_payload,
    build_equipment_timeline_payload,
    build_non_scheduled_filter_payload,
    build_non_scheduled_list_payload,
    build_non_scheduled_monthly_payload,
    build_non_scheduled_summary_payload,
)
from spare_parts_service import (
    build_spare_parts_payload,
    build_project_transactions_payload,
    build_all_years_transactions_payload,
    build_external_po_payload,
    import_spare_inventory_file,
    import_external_po_file,
    import_project_transactions_file,
    get_maintenance_import_status,
)
# downtime_service is needed by the Maintenance "Analysis" tab
# (/api/downtime?period=all_years&work_orders_only=1) and by spare_parts_service
from downtime_service import (
    build_downtime_payload,
    DOWNTIME_CACHE_OUTPUT_FILE,
    import_work_order_file,
    get_work_order_import_status,
    build_mtbf_work_order_history_payload,
)

# ── Path configuration ────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", "frontend"))
DATA_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", "data"))

# ── App setup ─────────────────────────────────────────────────────────────────
app = Flask(__name__, static_folder=FRONTEND_DIR)


@app.after_request
def apply_cache_headers(response):
    """Prevent stale caches for API JSON; allow short browser cache for static assets."""
    if request.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store"
    return response


# ── Frontend routes ───────────────────────────────────────────────────────────

@app.route("/")
def root():
    """Root URL serves the Maintenance page directly."""
    return send_from_directory(os.path.join(FRONTEND_DIR, "Maintenance"), "index.html")


@app.route("/<path:path>")
def frontend_files(path):
    """Catch-all static file server for CSS, JS, shared assets, etc."""
    return send_from_directory(FRONTEND_DIR, path)


# ── Asset list (shared by Maintenance Analysis tab and Downtime) ──────────────

@app.route("/api/asset-list")
def asset_list_api():
    import openpyxl
    candidates = [
        os.path.join(DATA_DIR, "AssetList.xlsx"),
        os.path.join(BASE_DIR, "AssetList.xlsx"),
    ]
    path = next((p for p in candidates if os.path.exists(p)), None)
    if not path:
        return jsonify({"machines": [], "error": "AssetList.xlsx not found"}), 404
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Machine Criticality"]
        machines = []
        for r in range(2, ws.max_row + 1):
            raw_asset_id = ws.cell(r, 1).value
            machine_name = str(ws.cell(r, 2).value or "").strip()
            location = str(ws.cell(r, 3).value or "").strip()
            criticality = str(ws.cell(r, 4).value or "").strip()
            if not machine_name:
                continue
            asset_lines = [line.strip() for line in str(raw_asset_id or "").split("\n") if line.strip()]
            assets = []
            for line in asset_lines:
                ids = re.findall(r"[A-Z]{2,}[A-Z0-9]*-\d+", line)
                label_part = re.split(r":\s*", line, maxsplit=1)
                label = label_part[0].strip() if len(label_part) > 1 else ""
                for aid in ids:
                    assets.append({"asset_id": aid, "label": label or aid})
            machines.append({
                "machine_name": machine_name,
                "location": location,
                "criticality": criticality,
                "asset_count": len(assets),
                "assets": assets,
            })
        return jsonify({"machines": machines})
    except Exception as exc:
        return jsonify({"machines": [], "error": str(exc)}), 500


# ── Downtime routes (needed by Maintenance "Analysis" tab) ───────────────────

@app.route("/api/downtime")
def downtime_data():
    period = request.args.get("period")
    month = request.args.get("month")
    work_orders_only = str(request.args.get("work_orders_only", "")).strip().lower() in {"1", "true", "yes", "on"}
    return jsonify(build_downtime_payload(period, month, request.args.get("start"), request.args.get("end"), work_orders_only=work_orders_only))


@app.route("/api/downtime/import-work-orders", methods=["GET", "POST"])
def downtime_import_work_orders():
    if request.method == "GET":
        return jsonify(get_work_order_import_status())
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"ok": False, "message": "No work order file uploaded."}), 400
    replace = str(request.form.get("replace", "true")).strip().lower() not in {"0", "false", "no"}
    result = import_work_order_file(upload, replace=replace)
    return jsonify(result), (200 if result.get("ok") else 400)


@app.route("/api/downtime/mtbf-history")
def downtime_mtbf_history():
    return jsonify(build_mtbf_work_order_history_payload())


# ── Maintenance API routes ────────────────────────────────────────────────────

@app.route("/api/maintenance/overview")
def maintenance_overview():
    return jsonify(
        build_maintenance_overview_payload(
            month_value=request.args.get("month"),
            status=request.args.get("status", "all"),
            category=request.args.get("category", "all"),
            search=request.args.get("search", ""),
            sort=request.args.get("sort", "date_asc"),
            year=request.args.get("year", type=int),
            mix_month_value=request.args.get("mix_month"),
        )
    )


@app.route("/api/maintenance/utility/summary")
def maintenance_utility_summary():
    return jsonify(build_summary_payload(request.args.get("year", type=int)))


@app.route("/api/maintenance/utility/monthly")
def maintenance_utility_monthly():
    return jsonify(build_monthly_payload(request.args.get("month"), request.args.get("year", type=int)))


@app.route("/api/maintenance/utility/list")
def maintenance_utility_list():
    return jsonify(
        build_list_payload(
            month_value=request.args.get("month"),
            status=request.args.get("status", "all"),
            category=request.args.get("category", "all"),
            location=request.args.get("location", "all"),
            inspection=request.args.get("inspection", "all"),
            search=request.args.get("search", ""),
            sort=request.args.get("sort", "due_date_asc"),
            year=request.args.get("year", type=int),
            aggregate=request.args.get("aggregate", "occurrence"),
        )
    )


@app.route("/api/maintenance/utility/timeline")
def maintenance_utility_timeline():
    return jsonify(build_timeline_payload(request.args.get("year", type=int), request.args.get("month")))


@app.route("/api/maintenance/utility/filters")
def maintenance_utility_filters():
    return jsonify(build_filter_payload(request.args.get("year", type=int)))


@app.route("/api/maintenance/equipment/summary")
def maintenance_equipment_summary():
    return jsonify(build_equipment_summary_payload(request.args.get("year", type=int)))


@app.route("/api/maintenance/equipment/monthly")
def maintenance_equipment_monthly():
    return jsonify(build_equipment_monthly_payload(request.args.get("month"), request.args.get("year", type=int)))


@app.route("/api/maintenance/equipment/list")
def maintenance_equipment_list():
    return jsonify(
        build_equipment_list_payload(
            month_value=request.args.get("month"),
            status=request.args.get("status", "all"),
            category=request.args.get("category", "all"),
            location=request.args.get("location", "all"),
            inspection=request.args.get("inspection", "all"),
            search=request.args.get("search", ""),
            sort=request.args.get("sort", "due_date_asc"),
            year=request.args.get("year", type=int),
            aggregate=request.args.get("aggregate", "occurrence"),
            priority=request.args.get("priority", "all"),
            critical=request.args.get("critical", "all"),
            week=request.args.get("week", "all"),
        )
    )


@app.route("/api/maintenance/equipment/timeline")
def maintenance_equipment_timeline():
    return jsonify(build_equipment_timeline_payload(request.args.get("year", type=int), request.args.get("month")))


@app.route("/api/maintenance/equipment/filters")
def maintenance_equipment_filters():
    return jsonify(build_equipment_filter_payload(request.args.get("year", type=int)))


@app.route("/api/maintenance/non_scheduled/summary")
def maintenance_non_scheduled_summary():
    return jsonify(build_non_scheduled_summary_payload(request.args.get("year", type=int)))


@app.route("/api/maintenance/non_scheduled/monthly")
def maintenance_non_scheduled_monthly():
    return jsonify(build_non_scheduled_monthly_payload(request.args.get("month"), request.args.get("year", type=int)))


@app.route("/api/maintenance/non_scheduled/list")
def maintenance_non_scheduled_list():
    return jsonify(
        build_non_scheduled_list_payload(
            month_value=request.args.get("month"),
            status=request.args.get("status", "all"),
            priority=request.args.get("priority", "all"),
            area=request.args.get("area", "all"),
            search=request.args.get("search", ""),
            year=request.args.get("year", type=int),
            sort=request.args.get("sort", "due_date_asc"),
        )
    )


@app.route("/api/maintenance/non_scheduled/filters")
def maintenance_non_scheduled_filters():
    return jsonify(build_non_scheduled_filter_payload(request.args.get("year", type=int)))


@app.route("/api/maintenance/spare_parts")
def maintenance_spare_parts():
    return jsonify(build_spare_parts_payload())


@app.route("/api/maintenance/project_transactions")
def maintenance_project_transactions():
    return jsonify(build_project_transactions_payload())


@app.route("/api/maintenance/project_transactions_all")
def maintenance_project_transactions_all():
    return jsonify(build_all_years_transactions_payload())


@app.route("/api/maintenance/external_po")
def maintenance_external_po():
    return jsonify(build_external_po_payload())


@app.route("/api/maintenance/import-status")
def maintenance_import_status():
    return jsonify(get_maintenance_import_status())


@app.route("/api/maintenance/import/spare-inventory", methods=["POST"])
def maintenance_import_spare_inventory():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"ok": False, "message": "No inventory file uploaded."}), 400
    result = import_spare_inventory_file(upload)
    return jsonify(result), (200 if result.get("ok") else 400)


@app.route("/api/maintenance/import/external-po", methods=["POST"])
def maintenance_import_external_po():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"ok": False, "message": "No external parts file uploaded."}), 400
    result = import_external_po_file(upload)
    return jsonify(result), (200 if result.get("ok") else 400)


@app.route("/api/maintenance/import/project-transactions", methods=["POST"])
def maintenance_import_project_transactions():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"ok": False, "message": "No project transactions file uploaded."}), 400
    result = import_project_transactions_file(upload)
    return jsonify(result), (200 if result.get("ok") else 400)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    debug = os.environ.get("FLASK_DEBUG", "0") not in {"0", "false", "no"}
    print(f"Maintenance standalone server starting on http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=debug)
