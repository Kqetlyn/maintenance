"""
Unified Preventive Maintenance (PM) schedule tracking service.

This module sits on top of the existing Stage 1 (utility) and Stage 2 (equipment)
PM schedule loaders in ``maintenance_service`` and normalises both into a single
``pmScheduleTasks`` dataset. Stage is derived purely from ``Asset_Master.xlsx``
(via ``asset_mapping``) — it is never hard-coded per source file.

The single public entry point ``build_pm_schedule_payload`` returns a management
ready payload for the Overview / Equipment / Utilities tabs, filtered by a global
Stage selector (All / Stage 1 / Stage 2).

Completion handling: the source schedules carry no explicit completion column, so
"completed" is INFERRED (a PM whose planned week is already in the past is treated
as done). This matches the existing dashboard behaviour, so PM Compliance % is a
real planning-based number rather than "N/A".
"""

from __future__ import annotations

import openpyxl
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from maintenance_service import DATA_DIR, build_equipment_dataset_from_path, build_utility_dataset
from pm_schedule_sources import (
    build_stage2_utility_dataset,
    get_pm_schedule_source_status,
    summarize_pm_schedule_sources,
)
from asset_mapping import load_asset_mapping
from pm_schedule_overrides import apply_overrides_and_autodone
from pm_planner_store import list_manual_tasks

# ── Functional-location mapping (Asset_Master: Asset Installation + Functional Locations) ──
_FL_CACHE = {"sig": None, "data": None}
UNMAPPED_FL = "Unmapped Functional Location"


def load_functional_location_map() -> dict:
    """Return {ASSET_ID: {code, name, type, parent, label}} resolved to the asset's
    ZONE-level functional location (e.g. ZN3 — Zone 3 Assembly). Cached by file sig."""
    path = Path(DATA_DIR) / "master" / "Asset_Master.xlsx"
    try:
        st = path.stat()
        sig = (st.st_mtime_ns, st.st_size)
    except OSError:
        return {}
    if _FL_CACHE["sig"] == sig and _FL_CACHE["data"] is not None:
        return _FL_CACHE["data"]

    fl_master, asset_fl = {}, {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Functional Locations" in wb.sheetnames:
            for i, row in enumerate(wb["Functional Locations"].iter_rows(values_only=True)):
                if i == 0 or not row or not row[0]:
                    continue
                code = str(row[0]).strip()
                # Normalise en/em dashes inside the name so "Zone 3 – Assembly"
                # reads cleanly as "Zone 3 Assembly" in the "Code – Name" label.
                raw_name = str(row[1] or "").replace("–", " ").replace("—", " ").replace("�", " ")
                fl_master[code] = {
                    "name": " ".join(raw_name.split()),
                    "type": str(row[2] or "").strip(),
                    "parent": str(row[3] or "").strip(),
                }
        if "Asset Installation" in wb.sheetnames:
            for i, row in enumerate(wb["Asset Installation"].iter_rows(values_only=True)):
                if i == 0 or not row or not row[0]:
                    continue
                asset = str(row[0]).strip().upper()
                fl = str(row[1] or "").strip()
                if asset and fl:
                    asset_fl[asset] = fl
        wb.close()
    except Exception:
        return {}

    def resolve_zone(code):
        seen, cur = set(), code
        while cur and cur in fl_master and cur not in seen:
            seen.add(cur)
            if fl_master[cur].get("type") == "ZONE":
                return cur
            cur = fl_master[cur].get("parent")
        head = code.split("-")[0] if code else ""
        if head in fl_master and fl_master[head].get("type") == "ZONE":
            return head
        return code if code in fl_master else None

    out = {}
    for asset, fl in asset_fl.items():
        resolved = resolve_zone(fl)
        if resolved and resolved in fl_master:
            entry = fl_master[resolved]
            out[asset] = {
                "code": resolved, "name": entry["name"], "type": entry.get("type"),
                "parent": entry.get("parent"), "label": f"{resolved} – {entry['name']}",
            }
    _FL_CACHE.update(sig=sig, data=out)
    return out


def _attach_functional_location(tasks: list[dict]) -> None:
    fl_map = load_functional_location_map()
    for task in tasks:
        fl = fl_map.get(str(task.get("assetId") or "").upper())
        if fl:
            task["functionalLocationCode"] = fl["code"]
            task["functionalLocationName"] = fl["name"]
            task["functionalLocationLabel"] = fl["label"]
        else:
            task["functionalLocationCode"] = ""
            task["functionalLocationName"] = UNMAPPED_FL
            task["functionalLocationLabel"] = UNMAPPED_FL

# ── Constants ──────────────────────────────────────────────────────────────────
MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

STAGE_VALUES = ("Stage 1", "Stage 2")
EQUIPMENT_GROUPS = {"Production Equipment"}

# Forward-looking window (in days) used for "Due Soon".
DUE_SOON_WINDOW_DAYS = 30

# Status categories
STATUS_NOT_DUE = "Not Due"
STATUS_DUE_THIS_MONTH = "Due This Month"
STATUS_DUE_SOON = "Due Soon"
STATUS_COMPLETED = "Completed"
STATUS_COMPLETED_LATE = "Completed Late"
STATUS_OVERDUE = "Overdue"
STATUS_MISSING_DATE = "Missing Schedule Date"
STATUS_MISSING_MAPPING = "Missing Asset Mapping"
STATUS_NEEDS_REVIEW = "Needs Review"

_FREQ_LABELS = {
    "monthly": "Monthly",
    "weekly": "Weekly",
    "quarterly": "Quarterly",
    "biweekly": "Bi-weekly",
    "bi-weekly": "Bi-weekly",
    "daily": "Daily",
    "annual": "Annual",
    "annually": "Annual",
    "yearly": "Annual",
    "semiannual": "Semi-annual",
    "semi-annual": "Semi-annual",
    "scheduled": "Scheduled",
}


# ── Helpers ────────────────────────────────────────────────────────────────────
def _normalize_stage_filter(value) -> str:
    text = str(value or "all").strip().lower().replace(" ", "")
    if text in {"stage1", "s1", "1"}:
        return "Stage 1"
    if text in {"stage2", "s2", "2"}:
        return "Stage 2"
    return "all"


def _quarter_label(month: int | None) -> str | None:
    if not month:
        return None
    return f"Q{(int(month) - 1) // 3 + 1}"


def _freq_label(occ) -> str:
    explicit_label = str(occ.get("frequency_label") or "").strip()
    if explicit_label:
        return explicit_label
    ft = str(occ.get("frequency_type") or "").strip().lower()
    if ft in _FREQ_LABELS:
        return _FREQ_LABELS[ft]
    return ft.title() if ft else ""


def _pm_description(occ) -> str:
    freq = _freq_label(occ)
    base = f"{freq} Preventive Maintenance".strip() if freq else "Preventive Maintenance"
    if occ.get("inspection_required"):
        return f"{base} + additional checks"
    return base


def _parse_iso_date(value) -> date | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)[:10]).date()
    except (ValueError, TypeError):
        return None


def _derive_inferred_schedule_state(occ, today: date) -> dict:
    planned_date = _parse_iso_date(occ.get("scheduled_date"))
    scheduled_week_end = _parse_iso_date(occ.get("scheduled_week_end"))
    if planned_date and scheduled_week_end is None:
        scheduled_week_end = planned_date + timedelta(days=6)

    explicit_done = occ.get("is_done")
    if explicit_done is None:
        is_done = bool(scheduled_week_end and scheduled_week_end < today)
    else:
        is_done = bool(explicit_done)

    explicit_due_this_month = occ.get("is_due_this_month")
    if explicit_due_this_month is None:
        is_due_this_month = bool(
            (not is_done)
            and planned_date is not None
            and planned_date.year == today.year
            and planned_date.month == today.month
        )
    else:
        is_due_this_month = bool(explicit_due_this_month)

    is_due_soon = bool(
        (not is_done)
        and planned_date is not None
        and today < planned_date <= today.fromordinal(today.toordinal() + DUE_SOON_WINDOW_DAYS)
    )

    explicit_overdue = occ.get("is_overdue")
    if explicit_overdue is None:
        is_overdue = bool(
            (not is_done)
            and scheduled_week_end is not None
            and scheduled_week_end < today
            and not is_due_this_month
        )
    else:
        is_overdue = bool(explicit_overdue)

    try:
        explicit_days_overdue = int(occ.get("days_overdue"))
    except (TypeError, ValueError):
        explicit_days_overdue = None
    if explicit_days_overdue is not None:
        days_overdue = explicit_days_overdue if is_overdue else 0
    elif is_overdue and scheduled_week_end is not None:
        days_overdue = max((today - scheduled_week_end).days, 0)
    else:
        days_overdue = 0

    return {
        "planned_date": planned_date,
        "is_done": is_done,
        "is_due_this_month": is_due_this_month,
        "is_due_soon": is_due_soon,
        "is_overdue": is_overdue,
        "days_overdue": days_overdue,
    }


def _scope_for_group(main_group: str | None) -> str:
    return "equipment" if (main_group or "") in EQUIPMENT_GROUPS else "utility"


_REFRIGERATION_KEYWORDS = (
    "refriger", "condens", "evapor", "chiller", "cold room", "freezer", "cooling", "ammonia",
)


def _derive_group_from_schedule(occ, domain):
    """Derive Main Asset Group / System-Area from the schedule itself when the
    Asset ID cannot be joined to Asset_Master (the utility schedule uses its own
    UL- code scheme that does not exist in the master)."""
    cat = str(occ.get("category") or "").strip()
    sub = str(occ.get("subcategory") or "").strip()
    loc = str(occ.get("location_display") or occ.get("location_raw") or "").strip()
    blob = " ".join([cat, sub, loc]).lower()
    if any(k in blob for k in _REFRIGERATION_KEYWORDS):
        main_group = "Refrigeration"
    elif domain == "equipment":
        main_group = "Production Equipment"
    else:
        main_group = "Utilities"
    if cat and cat.lower() != "others":
        system_area = cat
    else:
        system_area = loc or cat or "General"
    return main_group, system_area


# ── Normalisation ──────────────────────────────────────────────────────────────
def _normalize_occurrence(occ, *, domain, source_file, source_slot, source_label, default_stage, asset_map, today):
    asset_id_raw = str(occ.get("asset_code") or "").strip()
    asset_id = asset_id_raw.upper()
    mapping = asset_map.get(asset_id) if asset_id else None
    mapped_from_master = bool(mapping and mapping.get("mappedStage") in STAGE_VALUES)

    if mapped_from_master:
        # Asset ID matched the master and carries a valid stage.
        mapping_status = mapping.get("mappingStatus") or "Mapped"
        stage = mapping.get("mappedStage")
        asset_name = mapping.get("mappedAssetName") or occ.get("asset_name") or asset_id_raw
        main_group = mapping.get("mappedMainAssetGroup") or ""
        sub_group = mapping.get("mappedSubAssetGroup") or occ.get("subcategory") or ""
        system_area = mapping.get("mappedSystemArea") or ""
        location = mapping.get("mappedLocation") or occ.get("location_display") or occ.get("location_raw") or ""
    else:
        # No master match. Stage follows the source slot and groups come from the schedule.
        asset_name = occ.get("asset_name") or asset_id_raw
        main_group, system_area = _derive_group_from_schedule(occ, domain)
        sub_group = occ.get("subcategory") or ""
        location = occ.get("location_display") or occ.get("location_raw") or ""
        if not asset_id_raw:
            stage = "Needs Stage Review"
            mapping_status = "Missing Asset ID"
        elif source_slot in {"utility_stage1", "utility_stage2"}:
            stage = default_stage if default_stage in STAGE_VALUES else "Stage 1"
            mapping_status = "Schedule-defined"
        else:
            stage = default_stage if default_stage in STAGE_VALUES else "Stage 2"
            mapping_status = "Unmapped"

    schedule_state = _derive_inferred_schedule_state(occ, today)
    planned_date = schedule_state["planned_date"]
    planned_month = occ.get("planned_month") or (planned_date.month if planned_date else None)
    planned_year = planned_date.year if planned_date else None
    frequency = _freq_label(occ)
    is_done = schedule_state["is_done"]
    is_due_this_month = schedule_state["is_due_this_month"]
    is_due_soon = schedule_state["is_due_soon"]
    is_overdue = schedule_state["is_overdue"]

    # ── Schedule status (mapping issues take priority, then time-based) ──
    if mapping_status in {"Missing Asset ID", "Unmapped"}:
        schedule_status = STATUS_MISSING_MAPPING
    elif planned_date is None:
        schedule_status = STATUS_MISSING_DATE
    elif stage not in STAGE_VALUES:
        schedule_status = STATUS_NEEDS_REVIEW
    elif is_done:
        schedule_status = STATUS_COMPLETED
    elif is_due_this_month:
        schedule_status = STATUS_DUE_THIS_MONTH
    elif is_due_soon:
        schedule_status = STATUS_DUE_SOON
    elif is_overdue:
        schedule_status = STATUS_OVERDUE
    else:
        schedule_status = STATUS_NOT_DUE

    needs_review = (
        schedule_status in {STATUS_NEEDS_REVIEW, STATUS_MISSING_DATE}
        or stage not in STAGE_VALUES
        or not main_group
        or not frequency
    )

    return {
        "pmTaskId": f"{source_slot}-{asset_id or 'NOID'}-{occ.get('scheduled_date') or 'NODATE'}-{occ.get('source_sheet') or ''}",
        "stage": stage,
        "assetId": asset_id_raw,
        "assetName": asset_name,
        "mainAssetGroup": main_group or "Unmapped",
        "subAssetGroup": sub_group,
        "systemArea": system_area or "Unassigned",
        "location": location or "Unassigned",
        "pmDescription": _pm_description(occ),
        "frequency": frequency,
        "plannedYear": planned_year,
        "plannedMonth": planned_month,
        "plannedMonthLabel": (MONTH_LABELS[planned_month - 1] if planned_month else ""),
        "plannedQuarter": _quarter_label(planned_month),
        "plannedDate": occ.get("scheduled_date"),
        "plannedDateLabel": occ.get("scheduled_date_label"),
        "contractorOrPIC": occ.get("assigned_technician") or "",
        "provider": "",
        "scheduleStatus": schedule_status,
        "completionStatus": "Completed (inferred)" if is_done else "Open",
        "actualCompletionDate": None,
        "daysOverdue": schedule_state["days_overdue"],
        "sourceFile": source_file,
        "sourceSlot": source_slot,
        "sourceLabel": source_label or source_slot,
        "sourceSheet": occ.get("source_sheet") or "",
        "mappingStatus": mapping_status,
        "domain": domain,
        "scope": _scope_for_group(main_group),
        # internal booleans (kept for aggregation; harmless in JSON)
        "isDone": is_done,
        "isDueThisMonth": is_due_this_month,
        "isDueSoon": is_due_soon,
        "isOverdue": is_overdue,
        "needsReview": needs_review,
    }


def _build_tasks(year, today):
    """Normalise the active PM schedule sources into one list."""
    mapping = load_asset_mapping(str(DATA_DIR))
    asset_map = mapping.get("asset_map", {}) if mapping.get("available") else {}
    source_status = get_pm_schedule_source_status()
    utility_stage1_source = source_status["utility_stage1"]
    equipment_stage1_source = source_status["equipment_stage1"]
    utility_stage2_source = source_status["utility_stage2"]
    equipment_stage2_source = source_status["equipment_stage2"]

    def source_path(source_entry):
        path_value = source_entry.get("path")
        return Path(path_value) if path_value else None

    tasks = []
    utility_stage1 = build_utility_dataset(year)
    for occ in utility_stage1.get("occurrences", []):
        tasks.append(_normalize_occurrence(
            occ,
            domain="utility",
            source_file=utility_stage1_source.get("file_name") or "",
            source_slot="utility_stage1",
            source_label=utility_stage1_source.get("label"),
            default_stage=utility_stage1_source.get("default_stage"),
            asset_map=asset_map,
            today=today,
        ))

    equipment_stage1 = build_equipment_dataset_from_path(
        source_path(equipment_stage1_source),
        year,
        cache_key_prefix="equipment_stage1_dataset",
        source_cache_key="equipment_stage1_asset_source",
        disk_cache_path=None,
    )
    for occ in equipment_stage1.get("occurrences", []):
        tasks.append(_normalize_occurrence(
            occ,
            domain="equipment",
            source_file=equipment_stage1_source.get("file_name") or "",
            source_slot="equipment_stage1",
            source_label=equipment_stage1_source.get("label"),
            default_stage=equipment_stage1_source.get("default_stage"),
            asset_map=asset_map,
            today=today,
        ))

    utility_stage2 = build_stage2_utility_dataset(year)
    for occ in utility_stage2.get("occurrences", []):
        tasks.append(_normalize_occurrence(
            occ,
            domain="utility",
            source_file=utility_stage2_source.get("file_name") or "",
            source_slot="utility_stage2",
            source_label=utility_stage2_source.get("label"),
            default_stage=utility_stage2_source.get("default_stage"),
            asset_map=asset_map,
            today=today,
        ))

    equipment_stage2 = build_equipment_dataset_from_path(
        source_path(equipment_stage2_source),
        year,
        cache_key_prefix="equipment_stage2_dataset",
        source_cache_key="equipment_stage2_asset_source",
        disk_cache_path=None,
    )
    for occ in equipment_stage2.get("occurrences", []):
        tasks.append(_normalize_occurrence(
            occ,
            domain="equipment",
            source_file=equipment_stage2_source.get("file_name") or "",
            source_slot="equipment_stage2",
            source_label=equipment_stage2_source.get("label"),
            default_stage=equipment_stage2_source.get("default_stage"),
            asset_map=asset_map,
            today=today,
        ))

    tracked_counts = Counter(task.get("sourceSlot") for task in tasks if task.get("sourceSlot"))
    source_summary = summarize_pm_schedule_sources(source_status, tracked_counts)

    meta = {
        "utilityLastSynced": utility_stage1.get("meta", {}).get("last_synced"),
        "utilityStage1LastSynced": utility_stage1.get("meta", {}).get("last_synced"),
        "equipmentStage1LastSynced": equipment_stage1.get("meta", {}).get("last_synced"),
        "utilityStage2LastSynced": utility_stage2.get("meta", {}).get("last_synced"),
        "equipmentLastSynced": equipment_stage2.get("meta", {}).get("last_synced"),
        "equipmentStage2LastSynced": equipment_stage2.get("meta", {}).get("last_synced"),
        "utilitySource": utility_stage1_source.get("file_name"),
        "utilityStage1Source": utility_stage1_source.get("file_name"),
        "equipmentStage1Source": equipment_stage1_source.get("file_name"),
        "utilityStage2Source": utility_stage2_source.get("file_name"),
        "equipmentSource": equipment_stage2_source.get("file_name"),
        "equipmentStage2Source": equipment_stage2_source.get("file_name"),
        "scheduleSources": list(source_status.values()),
        "sourceSummary": source_summary,
        "assetMasterAvailable": mapping.get("available", False),
        "assetMasterSynced": mapping.get("last_synced"),
        "errors": (
            (utility_stage1.get("meta", {}).get("errors") or [])
            + (equipment_stage1.get("meta", {}).get("errors") or [])
            + (utility_stage2.get("meta", {}).get("errors") or [])
            + (equipment_stage2.get("meta", {}).get("errors") or [])
        ),
    }
    return tasks, asset_map, meta


# ── Aggregations ───────────────────────────────────────────────────────────────
def _public_task(task):
    """Strip internal-only boolean fields for the JSON table rows."""
    drop = {"isDone", "isDueThisMonth", "isDueSoon", "isOverdue", "needsReview", "domain"}
    public = {k: v for k, v in task.items() if k not in drop}
    public["pmCategory"] = "Equipment" if task.get("scope") == "equipment" else "Utility"
    public["completionDate"] = task.get("actualCompletionDate")
    public["remarks"] = task.get("remarks") or ""
    return public


def _counter_to_chart(counter, *, top=None, order=None):
    items = list(counter.items())
    if order is not None:
        items = [(k, counter.get(k, 0)) for k in order]
    else:
        items.sort(key=lambda kv: (-kv[1], str(kv[0])))
    if top:
        items = items[:top]
    return {"labels": [str(k) for k, _ in items], "data": [v for _, v in items]}


def _kpis(tasks, *, today, sel_year, sel_month, mapped_asset_total):
    total = len(tasks)
    completed = sum(1 for t in tasks if t["isDone"])
    due_this_month = sum(
        1 for t in tasks
        if t["plannedMonth"] == sel_month and t["plannedYear"] == sel_year
    )
    due_soon = sum(1 for t in tasks if t["isDueSoon"])
    overdue = sum(1 for t in tasks if t["isOverdue"])
    backlog = sum(
        1 for t in tasks
        if (not t["isDone"]) and (t["isOverdue"] or t["isDueThisMonth"])
    )
    missing_mapping = sum(1 for t in tasks if t["mappingStatus"] in {"Missing Asset ID", "Unmapped"})
    needs_review = sum(1 for t in tasks if t["needsReview"])

    assets_with_pm = {t["assetId"].upper() for t in tasks if t["assetId"] and t["mappingStatus"] not in {"Missing Asset ID", "Unmapped"}}
    # The utility schedule uses a different code scheme than the master, so the
    # numerator can exceed the master count for that stage — clamp to 100%.
    coverage_pct = round(min(len(assets_with_pm), mapped_asset_total) / mapped_asset_total * 100, 1) if mapped_asset_total else None

    compliance = round(completed / total * 100, 1) if total else None

    return {
        "totalScheduled": total,
        "dueThisMonth": due_this_month,
        "dueSoon": due_soon,
        "completed": completed,
        "compliancePct": compliance,            # number (completion inferred) or None
        "overdue": overdue,
        "backlog": backlog,
        "coverage": {
            "assetsWithPm": len(assets_with_pm),
            "totalMappedAssets": mapped_asset_total,
            "pct": coverage_pct,
        },
        "missingMapping": missing_mapping,
        "needsReview": needs_review,
    }


def _data_quality(tasks):
    seen = set()
    duplicates = 0
    for t in tasks:
        key = (t["assetId"].upper(), t["plannedDate"], t["pmDescription"])
        if key in seen:
            duplicates += 1
        else:
            seen.add(key)
    counts = {
        "missingAssetId": sum(1 for t in tasks if not t["assetId"]),
        "unmappedAssetId": sum(1 for t in tasks if t["mappingStatus"] == "Unmapped"),
        "missingPlanned": sum(1 for t in tasks if not t["plannedDate"]),
        "missingFrequency": sum(1 for t in tasks if not t["frequency"]),
        "missingGroup": sum(1 for t in tasks if t["mainAssetGroup"] in {"", "Unmapped", "Unknown / Review"}),
        "missingStage": sum(1 for t in tasks if t["stage"] not in STAGE_VALUES),
        "duplicates": duplicates,
        "needsReview": sum(1 for t in tasks if t["needsReview"]),
    }
    # Compact review rows (capped) for the Overview expandable table.
    rows = [
        _public_task(t) for t in tasks
        if t["needsReview"] or t["mappingStatus"] in {"Missing Asset ID", "Unmapped"}
    ][:200]
    return counts, rows


def _monthly_series(tasks, sel_year):
    scheduled = [0] * 12
    completed = [0] * 12
    overdue = [0] * 12
    for t in tasks:
        m = t["plannedMonth"]
        if not m or not (1 <= m <= 12):
            continue
        if t["plannedYear"] and t["plannedYear"] != sel_year:
            continue
        scheduled[m - 1] += 1
        if t["isDone"]:
            completed[m - 1] += 1
        if t["isOverdue"]:
            overdue[m - 1] += 1
    return {
        "labels": MONTH_LABELS,
        "scheduled": scheduled,
        "completed": completed,
        "overdue": overdue,
    }


def _stage_breakdown(tasks):
    counter = Counter(t["stage"] for t in tasks)
    order = ["Stage 1", "Stage 2"] + sorted(k for k in counter if k not in {"Stage 1", "Stage 2"})
    return _counter_to_chart(counter, order=[k for k in order if counter.get(k)])


def _overview_section(tasks, *, today, sel_year, sel_month, mapped_asset_total):
    kpis = _kpis(tasks, today=today, sel_year=sel_year, sel_month=sel_month,
                 mapped_asset_total=mapped_asset_total)
    dq_counts, dq_rows = _data_quality(tasks)

    overdue_tasks = [t for t in tasks if t["isOverdue"]]
    due_soon_tasks = [t for t in tasks if t["isDueSoon"]]

    charts = {
        "scheduledByStage": _stage_breakdown(tasks),
        "scheduledByMonth": _monthly_series(tasks, sel_year),
        "overdueByStage": _stage_breakdown(overdue_tasks),
        "dueSoonByStage": _stage_breakdown(due_soon_tasks),
        "workloadByMainGroup": _counter_to_chart(Counter(t["mainAssetGroup"] for t in tasks), top=10),
        "workloadBySystemArea": _counter_to_chart(
            Counter(t["systemArea"] for t in tasks if t["systemArea"] and t["systemArea"] != "Unassigned"),
            top=10,
        ),
        "stageWorkload": _counter_to_chart(
            Counter(t["stage"] for t in tasks),
            order=["Stage 1", "Stage 2"],
        ),
        "dataQuality": {
            "labels": ["Missing Asset ID", "Unmapped", "Missing Date", "Missing Freq.", "Missing Group", "Duplicates", "Needs Review"],
            "data": [
                dq_counts["missingAssetId"], dq_counts["unmappedAssetId"], dq_counts["missingPlanned"],
                dq_counts["missingFrequency"], dq_counts["missingGroup"], dq_counts["duplicates"],
                dq_counts["needsReview"],
            ],
        },
    }
    return {
        "kpis": kpis,
        "charts": charts,
        "dataQuality": {"counts": dq_counts, "rows": dq_rows},
    }


def _sort_table(rows):
    status_rank = {
        STATUS_OVERDUE: 0, STATUS_DUE_THIS_MONTH: 1, STATUS_DUE_SOON: 2,
        STATUS_NEEDS_REVIEW: 3, STATUS_MISSING_MAPPING: 3, STATUS_MISSING_DATE: 3,
        STATUS_NOT_DUE: 4, STATUS_COMPLETED: 5, STATUS_COMPLETED_LATE: 5,
    }
    return sorted(
        rows,
        key=lambda t: (status_rank.get(t["scheduleStatus"], 9), -(t["daysOverdue"] or 0), t["plannedDate"] or "9999"),
    )


def _scope_section(tasks, *, scope, today, sel_year, sel_month, mapped_asset_total, group_chart_key):
    scoped = [t for t in tasks if t["scope"] == scope]
    kpis = _kpis(scoped, today=today, sel_year=sel_year, sel_month=sel_month,
                 mapped_asset_total=mapped_asset_total)

    if group_chart_key == "systemArea":
        secondary = _counter_to_chart(
            Counter(t["systemArea"] for t in scoped if t["systemArea"] and t["systemArea"] != "Unassigned"),
            top=12,
        )
        top_label = "topSystems"
        top_counter = Counter(t["systemArea"] for t in scoped if t["systemArea"] and t["systemArea"] != "Unassigned")
    else:
        secondary = _counter_to_chart(Counter(t["subAssetGroup"] for t in scoped if t["subAssetGroup"]), top=12)
        top_label = "topAssets"
        top_counter = Counter(f"{t['assetId']} — {t['assetName']}" for t in scoped if t["assetId"])

    charts = {
        "byMainGroup": _counter_to_chart(Counter(t["mainAssetGroup"] for t in scoped), top=10),
        "bySecondary": secondary,
        "byMonth": _monthly_series(scoped, sel_year),
        "byStage": _stage_breakdown(scoped),
    }

    all_rows = [_public_task(t) for t in _sort_table(scoped)]
    overdue = [_public_task(t) for t in _sort_table([t for t in scoped if t["isOverdue"]])]
    due_soon = [_public_task(t) for t in _sort_table([t for t in scoped if t["isDueSoon"]])]
    needs_review = [_public_task(t) for t in scoped if t["needsReview"]]

    return {
        "kpis": kpis,
        "charts": charts,
        top_label: _counter_to_chart(top_counter, top=10),
        "tables": {
            "all": all_rows[:1000],
            "allCount": len(all_rows),
            "overdue": overdue[:300],
            "dueSoon": due_soon[:300],
            "needsReview": needs_review[:300],
        },
    }


def _schedule_filter_options(tasks) -> dict:
    def unique_values(key):
        return sorted({
            str(task.get(key) or "").strip()
            for task in tasks
            if str(task.get(key) or "").strip()
        }, key=str.lower)

    return {
        "categories": ["All", "Utility", "Equipment"],
        "systems": unique_values("systemArea"),
        "assetGroups": unique_values("mainAssetGroup"),
        "statuses": unique_values("scheduleStatus"),
        "pics": unique_values("contractorOrPIC"),
    }


def _schedule_section(tasks) -> dict:
    rows = [_public_task(t) for t in _sort_table(tasks)]
    overdue = [_public_task(t) for t in _sort_table([t for t in tasks if t["isOverdue"]])]
    due_soon = [_public_task(t) for t in _sort_table([t for t in tasks if t["isDueSoon"]])]
    needs_review = [_public_task(t) for t in tasks if t["needsReview"]]
    return {
        "tasks": rows,
        "filterOptions": _schedule_filter_options(tasks),
        "tables": {
            "all": rows,
            "allCount": len(rows),
            "overdue": overdue,
            "dueSoon": due_soon,
            "needsReview": needs_review,
        },
    }


def _count_mapped_assets(asset_map, *, stage, scope=None):
    total = 0
    for entry in asset_map.values():
        entry_stage = entry.get("mappedStage")
        if entry_stage not in STAGE_VALUES:
            continue
        if stage != "all" and entry_stage != stage:
            continue
        if scope is not None and _scope_for_group(entry.get("mappedMainAssetGroup")) != scope:
            continue
        total += 1
    return total


# ── Public entry point ─────────────────────────────────────────────────────────
def build_pm_schedule_payload(stage="all", year=None, month=None):
    today = datetime.now().date()
    sel_year = int(year) if year else today.year
    stage_filter = _normalize_stage_filter(stage)

    try:
        sel_month = int(month) if month else today.month
    except (TypeError, ValueError):
        sel_month = today.month
    if not (1 <= sel_month <= 12):
        sel_month = today.month

    all_tasks, asset_map, source_meta = _build_tasks(sel_year, today)

    # Merge local PM status overrides and apply the auto-assumed-done rule so the
    # operational `status` (Done / Auto Done / Backlog / Deferred / …) is available
    # to every KPI, chart, table, and calendar cell downstream.
    override_stats = apply_overrides_and_autodone(all_tasks, today)
    source_meta["overrideStats"] = override_stats

    # Merge manually planned PM tasks (saved in data/pm_planner_tasks.json). These
    # carry source="Manual" and are already normalised to the imported task shape.
    manual_tasks = list_manual_tasks(today)
    all_tasks.extend(manual_tasks)
    source_meta["manualCount"] = len(manual_tasks)

    # Attach functional-location (asset installation → zone) for the FL workload chart.
    _attach_functional_location(all_tasks)

    # Available years across both schedules (for the year selector).
    years = sorted({t["plannedYear"] for t in all_tasks if t["plannedYear"]})
    if sel_year not in years:
        years = sorted(set(years) | {sel_year})

    if stage_filter == "all":
        tasks = all_tasks
    else:
        tasks = [t for t in all_tasks if t["stage"] == stage_filter]

    mapped_total_all = _count_mapped_assets(asset_map, stage=stage_filter)
    mapped_total_equipment = _count_mapped_assets(asset_map, stage=stage_filter, scope="equipment")
    mapped_total_utility = _count_mapped_assets(asset_map, stage=stage_filter, scope="utility")

    overview = _overview_section(
        tasks, today=today, sel_year=sel_year, sel_month=sel_month,
        mapped_asset_total=mapped_total_all,
    )
    equipment = _scope_section(
        tasks, scope="equipment", today=today, sel_year=sel_year, sel_month=sel_month,
        mapped_asset_total=mapped_total_equipment, group_chart_key="subAssetGroup",
    )
    utility = _scope_section(
        tasks, scope="utility", today=today, sel_year=sel_year, sel_month=sel_month,
        mapped_asset_total=mapped_total_utility, group_chart_key="systemArea",
    )
    schedule = _schedule_section(tasks)

    return {
        "meta": {
            "stageFilter": stage_filter,
            "year": sel_year,
            "month": sel_month,
            "monthLabel": MONTH_LABELS[sel_month - 1],
            "availableYears": years,
            "availableStages": list(STAGE_VALUES),
            "today": today.isoformat(),
            "completionBasis": "inferred",
            "dueSoonWindowDays": DUE_SOON_WINDOW_DAYS,
            "taskCountAllStages": len(all_tasks),
            "taskCount": len(tasks),
            "generatedAt": datetime.now().isoformat(),
            **source_meta,
        },
        "overview": overview,
        "schedule": schedule,
        "equipment": equipment,
        "utility": utility,
    }
